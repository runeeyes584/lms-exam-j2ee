import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# Sample data
data = {
    'Type': [
        'MULTIPLE_CHOICE', 'MULTIPLE_CHOICE', 'TRUE_FALSE', 'TRUE_FALSE', 
        'FILL_IN', 'FILL_IN', 'MULTIPLE_CHOICE', 'MULTIPLE_CHOICE',
        'TRUE_FALSE', 'FILL_IN', 'MULTIPLE_CHOICE', 'TRUE_FALSE',
        'FILL_IN', 'MULTIPLE_CHOICE', 'TRUE_FALSE'
    ],
    'Content': [
        'What is the capital of France?',
        'Which of the following are programming languages? (Select all that apply)',
        'Java is a compiled language',
        'The Earth is flat',
        'The chemical symbol for gold is ___',
        'The capital of Vietnam is ___',
        'What is 2 + 2?',
        'Which HTTP methods are idempotent?',
        'MongoDB is a relational database',
        'In Java, the keyword to define a constant is ___',
        'What does REST stand for?',
        'Spring Boot requires XML configuration',
        'The default port for HTTP is ___',
        'Which design patterns are creational patterns?',
        'JSON stands for JavaScript Object Notation'
    ],
    'ImageUrl': [''] * 15,
    'Difficulty': [
        'RECOGNIZE', 'UNDERSTAND', 'UNDERSTAND', 'RECOGNIZE', 'RECOGNIZE', 'RECOGNIZE',
        'RECOGNIZE', 'ANALYZE', 'UNDERSTAND', 'APPLY', 'UNDERSTAND', 'RECOGNIZE',
        'RECOGNIZE', 'ANALYZE', 'RECOGNIZE'
    ],
    'Topics': [
        'Geography,Europe', 'Programming,Computer Science', 'Programming,Java',
        'Geography', 'Chemistry', 'Geography,Asia', 'Math,Arithmetic',
        'Web Development,REST API', 'Database,NoSQL', 'Programming,Java',
        'Web Development,API', 'Java,Spring Boot', 'Networking,Web',
        'Design Patterns,Software Engineering', 'Web Development,Data Format'
    ],
    'Points': [1, 2, 1, 1, 2, 1, 1, 3, 1.5, 2, 2, 1, 1, 3, 1],
    'Explanation': [
        'Paris is the capital and largest city of France',
        'Python and Java are programming languages',
        'Java compiles to bytecode which runs on JVM',
        'The Earth is an oblate spheroid',
        "Au comes from Latin 'aurum'",
        'Hanoi is the capital city',
        'Basic addition operation',
        'GET, PUT, DELETE are idempotent',
        'MongoDB is a NoSQL document database',
        'final keyword prevents reassignment',
        'REST is an architectural style',
        'Spring Boot uses annotation-based configuration by default',
        'HTTP uses port 80 by default',
        'Singleton and Factory are creational patterns',
        'JSON is a lightweight data interchange format'
    ],
    'Option1': [
        'London', 'Python', 'TRUE', 'FALSE', 'Au', 'Hanoi', '3', 'GET',
        'FALSE', 'final', 'Representational State Transfer', 'FALSE', '80',
        'Singleton', 'TRUE'
    ],
    'Option2': [
        'Paris', 'HTML', '', '', '', '', '4', 'POST', '', '', 
        'Really Simple Transfer', '', '', 'Observer', ''
    ],
    'Option3': [
        'Berlin', 'Java', '', '', '', '', '5', 'PUT', '', '',
        'Remote State Transaction', '', '', 'Factory', ''
    ],
    'Option4': [
        'Rome', 'CSS', '', '', '', '', '6', 'DELETE', '', '',
        'Resource Exchange System', '', '', 'Strategy', ''
    ],
    'IsCorrect1': [
        'FALSE', 'TRUE', '', '', '', '', 'FALSE', 'TRUE', '', '',
        'TRUE', '', '', 'TRUE', ''
    ],
    'IsCorrect2': [
        'TRUE', 'FALSE', '', '', '', '', 'TRUE', 'FALSE', '', '',
        'FALSE', '', '', 'FALSE', ''
    ],
    'IsCorrect3': [
        'FALSE', 'TRUE', '', '', '', '', 'FALSE', 'TRUE', '', '',
        'FALSE', '', '', 'TRUE', ''
    ],
    'IsCorrect4': [
        'FALSE', 'FALSE', '', '', '', '', 'FALSE', 'TRUE', '', '',
        'FALSE', '', '', 'FALSE', ''
    ]
}

# Create DataFrame
df = pd.DataFrame(data)

# Create Excel workbook
wb = Workbook()
ws = wb.active
ws.title = "Questions"

# Write header with formatting
header = list(df.columns)
ws.append(header)

# Style header row
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Write data rows
for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
    ws.append(row)
    
    # Color code by question type
    type_cell = ws.cell(row=r_idx, column=1)
    if type_cell.value == 'MULTIPLE_CHOICE':
        type_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    elif type_cell.value == 'TRUE_FALSE':
        type_cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    elif type_cell.value == 'FILL_IN':
        type_cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

# Adjust column widths
column_widths = {
    'A': 18, 'B': 60, 'C': 15, 'D': 12, 'E': 30, 'F': 8, 'G': 50,
    'H': 35, 'I': 35, 'J': 35, 'K': 35, 'L': 12, 'M': 12, 'N': 12, 'O': 12
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Add instructions sheet
ws_instructions = wb.create_sheet("Instructions")
instructions = [
    ["Excel Import Template - Quick Guide"],
    [""],
    ["Column Structure:"],
    ["A - Type: MULTIPLE_CHOICE, TRUE_FALSE, or FILL_IN"],
    ["B - Content: Question text (required)"],
    ["C - ImageUrl: Optional image URL"],
    ["D - Difficulty: RECOGNIZE, UNDERSTAND, APPLY, or ANALYZE (Bloom's Taxonomy)"],
    ["E - Topics: Comma-separated list"],
    ["F - Points: Numeric value (default: 1.0)"],
    ["G - Explanation: Answer explanation"],
    ["H-K - Options: For MULTIPLE_CHOICE (Option1-4), For TRUE_FALSE (H=TRUE/FALSE), For FILL_IN (H=answer)"],
    ["L-O - IsCorrect: TRUE/FALSE flags for each option (MULTIPLE_CHOICE only)"],
    [""],
    ["Import API:"],
    ["POST http://localhost:8080/api/questions/import/excel"],
    ["Content-Type: multipart/form-data"],
    ["Parameter: file (Excel .xlsx)"],
    [""],
    ["Tips:"],
    ["- Keep header row (row 1) unchanged"],
    ["- Leave optional cells empty"],
    ["- Use TRUE/FALSE for boolean values"],
    ["- Check EXCEL_IMPORT_GUIDE.md for details"]
]

for row in instructions:
    ws_instructions.append(row)

ws_instructions.column_dimensions['A'].width = 80

# Save workbook
filename = 'sample_questions.xlsx'
wb.save(filename)
print(f"✅ Created {filename} successfully!")
print(f"📊 Contains {len(df)} sample questions")
print(f"📝 Ready to import via POST /api/questions/import/excel")
