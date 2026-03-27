import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# Sample data (60 Java questions, balanced by type and difficulty)
DIFFICULTIES = ["RECOGNIZE", "UNDERSTAND", "APPLY", "ANALYZE"]
POINTS_BY_DIFFICULTY = {
    "RECOGNIZE": 1,
    "UNDERSTAND": 2,
    "APPLY": 3,
    "ANALYZE": 4,
}

MULTIPLE_CHOICE_QUESTIONS = [
    ("Which keyword declares a class in Java?", ["class", "struct", "type", "def"], 0, "Java uses the class keyword to declare classes.", "Java,Basics"),
    ("Which method is the entry point of a Java application?", ["start()", "run()", "main()", "init()"], 2, "The JVM starts from main(String[] args).", "Java,Basics"),
    ("Which primitive type stores 32-bit integers in Java?", ["short", "int", "long", "byte"], 1, "int is the 32-bit signed integer type.", "Java,Data Types"),
    ("Which keyword is used to inherit a class in Java?", ["implements", "extends", "inherits", "super"], 1, "A subclass extends a superclass with the extends keyword.", "Java,OOP"),
    ("Which package contains ArrayList?", ["java.util", "java.io", "java.lang", "java.net"], 0, "ArrayList is in the java.util package.", "Java,Collections"),
    ("What is the main purpose of encapsulation in Java?", ["To expose fields directly", "To hide implementation details", "To avoid classes", "To remove constructors"], 1, "Encapsulation restricts direct access and protects object state.", "Java,OOP"),
    ("Why does Java use a garbage collector?", ["To increase syntax complexity", "To manage memory automatically", "To replace the compiler", "To disable inheritance"], 1, "GC reclaims unreachable objects automatically.", "Java,JVM"),
    ("What does the final keyword do on a method?", ["Makes it static", "Makes it abstract", "Prevents overriding", "Prevents overloading"], 2, "A final method cannot be overridden in subclasses.", "Java,OOP"),
    ("When should you use an interface in Java?", ["When no contract is needed", "When defining a behavior contract", "When avoiding polymorphism", "When making all fields mutable"], 1, "Interfaces are used to define common contracts.", "Java,OOP"),
    ("What is the purpose of try-with-resources?", ["To skip exception handling", "To auto-close resources", "To speed up loops", "To create threads"], 1, "Resources implementing AutoCloseable are closed automatically.", "Java,Exceptions"),
    ("Given int[] a = {1,2,3}; what is a.length?", ["2", "3", "4", "Compilation error"], 1, "Array length is the number of elements.", "Java,Arrays"),
    ("What is printed by System.out.println(\"A\" + 5 + 2)?", ["A7", "A52", "7A", "Compilation error"], 1, "String concatenation is evaluated left to right.", "Java,Operators"),
    ("Which collection keeps insertion order and allows duplicates?", ["HashSet", "TreeSet", "ArrayList", "HashMap"], 2, "ArrayList preserves insertion order and allows duplicates.", "Java,Collections"),
    ("If a method throws IOException, which keyword is used in the signature?", ["catch", "final", "throws", "throwing"], 2, "Checked exceptions are declared with throws.", "Java,Exceptions"),
    ("Which loop is best when index-based access is required for a List?", ["Enhanced for", "while(true)", "for with index", "switch"], 2, "Index-based for loop gives direct index access.", "Java,Control Flow"),
    ("Which statement about HashMap and TreeMap is correct?", ["Both keep insertion order", "HashMap sorts keys naturally", "TreeMap keeps keys sorted", "Neither stores key-value pairs"], 2, "TreeMap maintains natural/comparator-based key ordering.", "Java,Collections"),
    ("Which design principle is applied by programming to interfaces?", ["Tight coupling", "High mutability", "Abstraction and loose coupling", "No polymorphism"], 2, "Coding to interfaces reduces coupling and improves flexibility.", "Java,Design Principles"),
    ("Which scenario can cause ConcurrentModificationException?", ["Modifying a list during for-each iteration", "Reading from an empty list", "Calling size() repeatedly", "Using final variables"], 0, "Structural changes during iteration trigger fail-fast behavior.", "Java,Collections"),
    ("What is the most suitable return type for Optional-based null handling?", ["void", "Object", "Optional<T>", "Throwable"], 2, "Optional<T> expresses an optional value explicitly.", "Java,Best Practices"),
    ("Which Java 8 feature allows passing behavior as data?", ["Packages", "Lambda expressions", "Annotations", "Enums"], 1, "Lambdas enable functional-style behavior passing.", "Java,Functional Programming"),
]

TRUE_FALSE_QUESTIONS = [
    ("Java source files are compiled into bytecode files with the .class extension.", True, "javac compiles .java into .class bytecode.", "Java,Compilation"),
    ("The java.lang package must always be imported explicitly.", False, "java.lang is imported automatically.", "Java,Basics"),
    ("A constructor can have a return type in Java.", False, "Constructors do not define return types.", "Java,OOP"),
    ("An abstract class can contain both abstract and concrete methods.", True, "Abstract classes can mix both method types.", "Java,OOP"),
    ("String objects are mutable in Java.", False, "String is immutable; operations return new objects.", "Java,Core API"),
    ("Overloading methods depends on different parameter lists.", True, "Method overloading uses same name with different params.", "Java,OOP"),
    ("Method overriding requires the same method signature in subclass.", True, "Overriding matches signature and compatible return type.", "Java,OOP"),
    ("A static method can access instance fields directly without an object.", False, "Static context cannot access instance members directly.", "Java,Static"),
    ("The finally block always executes, even after System.exit(0).", False, "System.exit may terminate the JVM before finally runs.", "Java,Exceptions"),
    ("A checked exception must be handled or declared.", True, "Compiler enforces checked exception handling.", "Java,Exceptions"),
    ("ArrayList is synchronized by default.", False, "ArrayList is not thread-safe by default.", "Java,Collections"),
    ("HashSet allows duplicate elements.", False, "Set implementations do not allow duplicates.", "Java,Collections"),
    ("TreeSet stores elements in sorted order.", True, "TreeSet is backed by a sorted tree structure.", "Java,Collections"),
    ("The volatile keyword guarantees atomicity for increment operations.", False, "volatile gives visibility, not atomic compound operations.", "Java,Concurrency"),
    ("synchronized can be applied to methods and code blocks.", True, "Both synchronized methods and blocks are valid.", "Java,Concurrency"),
    ("A deadlock can happen when threads wait on each other's locks.", True, "Circular wait is a classic deadlock condition.", "Java,Concurrency"),
    ("JVM, JRE, and JDK are exactly the same thing.", False, "JDK includes tools; JRE provides runtime; JVM executes bytecode.", "Java,Platform"),
    ("A local variable in Java gets a default value automatically.", False, "Local variables must be initialized before use.", "Java,Basics"),
    ("The equals method should be consistent with hashCode.", True, "Contract requires equal objects to have same hash code.", "Java,Core API"),
    ("Streams in Java 8 can be reused after a terminal operation.", False, "A stream is consumed after a terminal operation.", "Java,Streams"),
]

FILL_IN_QUESTIONS = [
    ("In Java, the keyword used to inherit from a class is ___.", "extends", "A subclass uses extends to inherit.", "Java,OOP"),
    ("The default value of a boolean instance field is ___.", "false", "Instance boolean fields default to false.", "Java,Basics"),
    ("The interface that allows lambda expressions with one abstract method is called a ___ interface.", "functional", "A functional interface has exactly one abstract method.", "Java,Functional Programming"),
    ("The method used to compare string content is ___.", "equals", "Use equals for content comparison, not ==.", "Java,Core API"),
    ("The package automatically imported in every Java file is ___.", "java.lang", "java.lang is implicitly available.", "Java,Basics"),
    ("To convert a String to int, you can use Integer.___(\"123\").", "parseInt", "Integer.parseInt converts numeric strings to int.", "Java,Core API"),
    ("To prevent a class from being inherited, mark it as ___.", "final", "final class cannot be subclassed.", "Java,OOP"),
    ("In a try-with-resources block, resources must implement ___.", "AutoCloseable", "AutoCloseable enables automatic closing.", "Java,Exceptions"),
    ("The keyword used to refer to the current object instance is ___.", "this", "this points to the current object.", "Java,OOP"),
    ("The immediate parent class of all Java classes is ___.", "Object", "java.lang.Object is the root class.", "Java,OOP"),
    ("To create an immutable list in Java 9+, use List.___(...).", "of", "List.of creates an unmodifiable list.", "Java,Collections"),
    ("The utility class for optional values without null is ___.", "Optional", "Optional models presence/absence of a value.", "Java,Best Practices"),
    ("In the expression i++, ++ is the ___ operator.", "increment", "++ increases a numeric value by one.", "Java,Operators"),
    ("The method that starts a thread is ___.", "start", "Thread.start creates a new call stack and invokes run.", "Java,Concurrency"),
    ("The collection interface that stores unique elements is ___.", "Set", "Set disallows duplicate elements.", "Java,Collections"),
    ("To sort a list with custom logic, pass a ___ to sort().", "Comparator", "Comparator defines custom ordering rules.", "Java,Collections"),
    ("The exception thrown when dividing an integer by zero is ___.", "ArithmeticException", "Integer division by zero raises ArithmeticException.", "Java,Exceptions"),
    ("The Java API for date and time introduced in Java 8 is java.time.___.", "package", "The modern date-time API is in java.time package.", "Java,Date Time"),
    ("The method reference operator in Java is ___.", "::", "Double colon creates a method reference.", "Java,Functional Programming"),
    ("To wait for a thread to finish, call thread.___().", "join", "join blocks until the target thread completes.", "Java,Concurrency"),
]


def difficulty_for_index(index):
    return DIFFICULTIES[index // 5]


rows = []

for index, (content, options, correct_idx, explanation, topics) in enumerate(MULTIPLE_CHOICE_QUESTIONS):
    difficulty = difficulty_for_index(index)
    rows.append({
        "Type": "MULTIPLE_CHOICE",
        "Content": content,
        "ImageUrl": "",
        "Difficulty": difficulty,
        "Topics": topics,
        "Points": POINTS_BY_DIFFICULTY[difficulty],
        "Explanation": explanation,
        "Option1": options[0],
        "Option2": options[1],
        "Option3": options[2],
        "Option4": options[3],
        "IsCorrect1": "TRUE" if correct_idx == 0 else "FALSE",
        "IsCorrect2": "TRUE" if correct_idx == 1 else "FALSE",
        "IsCorrect3": "TRUE" if correct_idx == 2 else "FALSE",
        "IsCorrect4": "TRUE" if correct_idx == 3 else "FALSE",
    })

for index, (statement, is_true, explanation, topics) in enumerate(TRUE_FALSE_QUESTIONS):
    difficulty = difficulty_for_index(index)
    rows.append({
        "Type": "TRUE_FALSE",
        "Content": statement,
        "ImageUrl": "",
        "Difficulty": difficulty,
        "Topics": topics,
        "Points": POINTS_BY_DIFFICULTY[difficulty],
        "Explanation": explanation,
        "Option1": "TRUE",
        "Option2": "FALSE",
        "Option3": "",
        "Option4": "",
        "IsCorrect1": "TRUE" if is_true else "FALSE",
        "IsCorrect2": "FALSE" if is_true else "TRUE",
        "IsCorrect3": "",
        "IsCorrect4": "",
    })

for index, (content, answer, explanation, topics) in enumerate(FILL_IN_QUESTIONS):
    difficulty = difficulty_for_index(index)
    rows.append({
        "Type": "FILL_IN",
        "Content": content,
        "ImageUrl": "",
        "Difficulty": difficulty,
        "Topics": topics,
        "Points": POINTS_BY_DIFFICULTY[difficulty],
        "Explanation": explanation,
        "Option1": answer,
        "Option2": "",
        "Option3": "",
        "Option4": "",
        "IsCorrect1": "",
        "IsCorrect2": "",
        "IsCorrect3": "",
        "IsCorrect4": "",
    })

# Create DataFrame
df = pd.DataFrame(rows)

# Create Excel workbook
wb = Workbook()
ws = wb.active
ws.title = "Questions"

# Write header with formatting
header = list(df.columns)
ws.append(header)

# Style header row
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Write data rows
for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
    ws.append(row)
    
    # Color code by question type
    type_cell = ws.cell(row=r_idx, column=1)
    if type_cell.value == 'MULTIPLE_CHOICE':
        type_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    elif type_cell.value == 'TRUE_FALSE':
        type_cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    elif type_cell.value == 'FILL_IN':
        type_cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

# Adjust column widths
column_widths = {
    'A': 18, 'B': 60, 'C': 15, 'D': 12, 'E': 30, 'F': 8, 'G': 50,
    'H': 35, 'I': 35, 'J': 35, 'K': 35, 'L': 12, 'M': 12, 'N': 12, 'O': 12
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Add instructions sheet
ws_instructions = wb.create_sheet("Instructions")
instructions = [
    ["Excel Import Template - Quick Guide"],
    [""],
    ["Column Structure:"],
    ["A - Type: MULTIPLE_CHOICE, TRUE_FALSE, or FILL_IN"],
    ["B - Content: Question text (required)"],
    ["C - ImageUrl: Optional image URL"],
    ["D - Difficulty: RECOGNIZE, UNDERSTAND, APPLY, or ANALYZE (Bloom's Taxonomy)"],
    ["E - Topics: Comma-separated list"],
    ["F - Points: Numeric value (default: 1.0)"],
    ["G - Explanation: Answer explanation"],
    ["H-K - Options: For MULTIPLE_CHOICE (Option1-4), For TRUE_FALSE (H=TRUE/FALSE), For FILL_IN (H=answer)"],
    ["L-O - IsCorrect: TRUE/FALSE flags for each option (MULTIPLE_CHOICE only)"],
    [""],
    ["Import API:"],
    ["POST http://localhost:8080/api/questions/import/excel"],
    ["Content-Type: multipart/form-data"],
    ["Parameter: file (Excel .xlsx)"],
    [""],
    ["Tips:"],
    ["- Keep header row (row 1) unchanged"],
    ["- Leave optional cells empty"],
    ["- Use TRUE/FALSE for boolean values"],
    ["- Check EXCEL_IMPORT_GUIDE.md for details"]
]

for row in instructions:
    ws_instructions.append(row)

ws_instructions.column_dimensions['A'].width = 80

# Save workbook
filename = 'sample_questions.xlsx'
wb.save(filename)
print(f"Created {filename} successfully.")
print(f"Contains {len(df)} sample questions.")
print("Ready to import via POST /api/questions/import/excel.")
